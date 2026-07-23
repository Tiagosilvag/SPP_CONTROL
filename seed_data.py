"""
Importa os dados da planilha original (Controle_Materiais_SPP.xlsx) para o
banco de dados do SPP - Control, preservando os IDs originais para manter a
integridade referencial entre as abas.

Uso:
    python seed_data.py
"""
import os
import sqlite3
import openpyxl

from config import Config

XLSX_PATH = os.path.join(os.path.dirname(__file__), "data", "Controle_Materiais_SPP.xlsx")
SCHEMA_PATH = os.path.join(os.path.dirname(__file__), "schema.sql")


def bool_sn(v):
    return 1 if str(v).strip().lower() == "sim" else 0


def as_date(v):
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date().isoformat()
    return str(v)


def run():
    db_path = Config.DATABASE_PATH
    os.makedirs(os.path.dirname(db_path), exist_ok=True)

    # Recria o banco do zero
    if os.path.exists(db_path):
        os.remove(db_path)

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")
    with open(SCHEMA_PATH, encoding="utf-8") as f:
        conn.executescript(f.read())

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    cur = conn.cursor()

    # --- Secretarias -----------------------------------------------------
    ws = wb["Secretarias"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cur.execute(
            "INSERT INTO secretarias (id, nome, sigla, responsavel, contato, ativa) VALUES (?,?,?,?,?,?)",
            (row[0], row[1], row[2], row[3], row[4], bool_sn(row[5])),
        )

    # --- Unidades ----------------------------------------------------------
    ws = wb["Unidades"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cur.execute(
            "INSERT INTO unidades (id, secretaria_id, nome, endereco, responsavel_local, ativa) VALUES (?,?,?,?,?,?)",
            (row[0], row[1], row[2], row[3], row[4], bool_sn(row[5])),
        )

    # --- Materiais -----------------------------------------------------------
    ws = wb["Materiais"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cur.execute(
            """INSERT INTO materiais (id, nome, tipo_material, categoria, unidade_medida, estoque_minimo, ativo)
               VALUES (?,?,?,?,?,?,?)""",
            (row[0], row[1], row[2], row[3], row[4], row[5] or 0, bool_sn(row[6])),
        )

    # --- Fornecedores -----------------------------------------------------------
    ws = wb["Fornecedores"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cur.execute(
            "INSERT INTO fornecedores (id, nome, cnpj, contato, ativo) VALUES (?,?,?,?,?)",
            (row[0], row[1], row[2], row[3], bool_sn(row[4])),
        )

    # --- Obras -----------------------------------------------------------------
    ws = wb["Obras"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cur.execute(
            """INSERT INTO obras (id, descricao, secretaria_solicitante_id, unidade_local_id, data_inicio,
               previsao_termino, status, fiscal_responsavel, observacoes) VALUES (?,?,?,?,?,?,?,?,?)""",
            (row[0], row[1], row[2], row[3], as_date(row[4]), as_date(row[5]), row[6], row[7], row[8]),
        )

    # --- Patrimônio --------------------------------------------------------------
    ws = wb["Patrimônio"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cur.execute(
            """INSERT INTO patrimonios (id, num_patrimonio, material_id, secretaria_proprietaria_id,
               data_aquisicao, estado_conservacao, unidade_atual_id, status, observacoes)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (
                row[0], row[1], row[2], row[4], as_date(row[5]), row[6], row[7], row[8], row[9],
            ),
        )

    # --- Entradas de Estoque -------------------------------------------------------
    ws = wb["Entradas_Estoque"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cur.execute(
            """INSERT INTO entradas_estoque (id, data_entrada, material_id, secretaria_proprietaria_id,
               unidade_destino_id, fornecedor_id, quantidade, nota_fiscal, obra_id, observacoes)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (
                row[0], as_date(row[1]), row[2], row[4], row[5], row[6], row[7], row[8], row[9], row[10],
            ),
        )

    # --- Movimentações de Consumíveis ----------------------------------------------
    ws = wb["Movimentacoes_Consumiveis"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cur.execute(
            """INSERT INTO movimentacoes_consumiveis (id, data_movimentacao, tipo_movimentacao, material_id,
               quantidade, unidade_origem_id, unidade_destino_id, secretaria_proprietaria_id, obra_id,
               data_prev_devolucao, data_real_devolucao, status_devolucao, responsavel, observacoes)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                row[0], as_date(row[1]), row[2], row[3], row[5], row[6], row[7], row[8], row[9],
                as_date(row[10]), as_date(row[11]), row[12] or "—", row[13], row[14],
            ),
        )

    # --- Movimentações de Patrimônio -------------------------------------------------
    ws = wb["Movimentacoes_Patrimônio"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cur.execute(
            """INSERT INTO movimentacoes_patrimonio (id, data_movimentacao, tipo_movimentacao, patrimonio_id,
               secretaria_proprietaria_id, unidade_origem_id, unidade_destino_id, obra_id,
               data_prev_devolucao, data_real_devolucao, status_devolucao, responsavel, observacoes)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                row[0], as_date(row[1]), row[2], row[3], row[5], row[6], row[7], row[8],
                as_date(row[9]), as_date(row[10]), row[11] or "—", row[12], row[13],
            ),
        )

    conn.commit()

    def count(table):
        return cur.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]

    print("Importação concluída com sucesso.")
    for t in [
        "secretarias", "unidades", "materiais", "fornecedores", "obras",
        "patrimonios", "entradas_estoque", "movimentacoes_consumiveis", "movimentacoes_patrimonio",
    ]:
        print(f"  {t}: {count(t)}")

    conn.close()


if __name__ == "__main__":
    run()
