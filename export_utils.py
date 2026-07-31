"""
Helpers genéricos de exportação usados pelo módulo de Relatórios: uma
lista de colunas (chave, rótulo) e uma lista de linhas (dicts) viram um
arquivo .xlsx (openpyxl) ou .pdf (reportlab), sem repetir código de
formatação em cada relatório.
"""
import io
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def _valor_texto(valor):
    if valor is None:
        return "-"
    return str(valor)


def exportar_excel(titulo, colunas, linhas):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31] if titulo else "Relatório"

    header_fill = PatternFill(start_color="EEF0FF", end_color="EEF0FF", fill_type="solid")
    for col_idx, (_, label) in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row_idx, linha in enumerate(linhas, start=2):
        for col_idx, (key, _) in enumerate(colunas, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_valor_texto(linha.get(key)))

    for col_idx, (_, label) in enumerate(colunas, start=1):
        largura = max(12, min(40, len(label) + 4))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{titulo or 'relatorio'}.xlsx".replace(" ", "_")
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def exportar_pdf(titulo, colunas, linhas):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elementos = [Paragraph(titulo or "Relatório", styles["Title"]), Spacer(1, 12)]

    cabecalho = [label for _, label in colunas]
    dados = [cabecalho] + [
        [_valor_texto(linha.get(key)) for key, _ in colunas] for linha in linhas
    ]

    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#5b5bf5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e6e8ef")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fb")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elementos.append(tabela)
    doc.build(elementos)
    buffer.seek(0)
    filename = f"{titulo or 'relatorio'}.pdf".replace(" ", "_")
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")
